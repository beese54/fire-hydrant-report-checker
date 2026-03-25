import streamlit as st

# ── Page config — must be the FIRST Streamlit call ────────────────────────────
st.set_page_config(
    page_title="SCDF Hydrant Form Checking",
    page_icon="🚒",
    layout="wide",
)

import os
import re
import json
import email
import email.policy
import io
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image

# ── Config ────────────────────────────────────────────────────────────────────
APP_PASSWORD = os.environ.get("APP_PASSWORD", "suffry")
MAX_UPLOAD_MB = int(os.environ.get("MAX_UPLOAD_MB", "20"))


# ─────────────────────────────────────────────────────────────────────────────
# EML PARSING
# ─────────────────────────────────────────────────────────────────────────────

def _extract_json_from_text(text: str):
    """Find a JSON array of {question, answer} objects embedded in arbitrary text."""
    # Try whole text first (body might be pure JSON)
    try:
        data = json.loads(text.strip())
        if isinstance(data, list) and data and "question" in data[0]:
            return data
    except Exception:
        pass

    # Scan for JSON arrays inside larger text (e.g. HTML body)
    for match in re.finditer(r'\[', text):
        start = match.start()
        # Walk forward to find matching ]
        depth = 0
        for i, ch in enumerate(text[start:]):
            if ch == '[':
                depth += 1
            elif ch == ']':
                depth -= 1
                if depth == 0:
                    candidate = text[start: start + i + 1]
                    try:
                        data = json.loads(candidate)
                        if isinstance(data, list) and data and "question" in data[0]:
                            return data
                    except Exception:
                        pass
                    break
    return None


def _parse_form_fields(data: list) -> dict:
    """Map FormSG question/answer pairs to named fields."""
    q_map = {
        item["question"].strip(): item.get("answer", "")
        for item in data
        if "question" in item
    }

    submission_id = q_map.get("Response ID", "").strip()
    fire_station = q_map.get("Select Fire Station", "").strip()
    defects = q_map.get("List of Defects", "").strip()

    # Hydrant table: answer is "street,type,number"
    street_name = hydrant_type = hydrant_number = ""
    for key, val in q_map.items():
        if "[table]" in key and "Hydrant" in key:
            parts = [p.strip() for p in val.split(",")]
            if len(parts) >= 3:
                street_name, hydrant_type, hydrant_number = parts[0], parts[1], parts[2]
            elif len(parts) == 2:
                street_name, hydrant_type = parts[0], parts[1]
            elif len(parts) == 1:
                street_name = parts[0]
            break

    # Photo filenames from "[attachment] Photos of Defect"
    photo_filenames = []
    for key, val in q_map.items():
        if "[attachment]" in key and "Photos of Defect" in key:
            photo_filenames = [f.strip() for f in val.split(",") if f.strip()]
            break

    return {
        "submission_id": submission_id,
        "fire_station": fire_station,
        "street_name": street_name,
        "hydrant_type": hydrant_type,
        "hydrant_number": hydrant_number,
        "defects": defects,
        "_photo_filenames": photo_filenames,
    }


def parse_eml(eml_bytes: bytes) -> dict:
    """
    Parse a FormSG .eml file.
    Returns a dict with submission fields + photos: [(filename, bytes), ...].
    Returns empty dict if JSON cannot be found.
    """
    msg = email.message_from_bytes(eml_bytes, policy=email.policy.compat32)

    form_data = {}
    attachments: dict[str, bytes] = {}

    for part in msg.walk():
        content_type = part.get_content_type()
        filename = part.get_filename()

        # Collect all named attachments (PIL will reject non-images at display time)
        if filename:
            payload = part.get_payload(decode=True)
            if payload:
                attachments[filename] = payload
            continue

        # Try to extract JSON from text parts
        if not form_data and content_type in ("text/plain", "text/html", "application/json"):
            try:
                payload = part.get_payload(decode=True)
                if payload:
                    text = payload.decode("utf-8", errors="ignore")
                    data = _extract_json_from_text(text)
                    if data:
                        form_data = _parse_form_fields(data)
            except Exception:
                pass

    if not form_data:
        return {}

    # Match photo filenames to attachment bytes (case-insensitive fallback)
    photos: list[tuple[str, bytes]] = []
    lower_attachments = {k.lower(): (k, v) for k, v in attachments.items()}
    for fname in form_data.pop("_photo_filenames", []):
        if fname in attachments:
            photos.append((fname, attachments[fname]))
        elif fname.lower() in lower_attachments:
            orig_name, data = lower_attachments[fname.lower()]
            photos.append((orig_name, data))

    # Fallback: if no named photos matched but attachments exist, show all of them
    if not photos and attachments:
        photos = list(attachments.items())
    form_data["photos"] = photos
    return form_data


def parse_contractor_eml(eml_bytes: bytes) -> dict:
    """
    Parse a contractor rectification FormSG .eml file.
    Returns a dict with submission_id, hydrant_number, timestamp,
    and labeled_photos: [(label, img_bytes), ...].
    Returns empty dict if no submission_id found.
    """
    msg = email.message_from_bytes(eml_bytes, policy=email.policy.compat32)

    attachments: dict[str, bytes] = {}
    raw_data = None

    for part in msg.walk():
        content_type = part.get_content_type()
        filename = part.get_filename()

        if content_type.startswith("image/") and filename:
            payload = part.get_payload(decode=True)
            if payload:
                attachments[filename] = payload
            continue

        if raw_data is None and content_type in ("text/plain", "text/html", "application/json"):
            try:
                payload = part.get_payload(decode=True)
                if payload:
                    text = payload.decode("utf-8", errors="ignore")
                    data = _extract_json_from_text(text)
                    if data:
                        raw_data = data
            except Exception:
                pass

    if not raw_data:
        return {}

    q_map = {
        item["question"].strip(): item.get("answer", "")
        for item in raw_data
        if "question" in item
    }

    submission_id = q_map.get("Submission ID", "").strip()
    hydrant_number = q_map.get("Hydrant Number", "").strip()
    timestamp = q_map.get("Timestamp", "").strip()

    if not submission_id:
        return {}

    # Ordered view labels mapped to their question keys
    view_keys = [
        ("[attachment] Front View of Hydrant", "Front View"),
        ("[attachment] Side View (Left) of Hydrant", "Left Side"),
        ("[attachment] Side View (Right) of Hydrant", "Right Side"),
        ("[attachment] Back View of Hydrant", "Back View"),
    ]

    lower_attachments = {k.lower(): (k, v) for k, v in attachments.items()}
    labeled_photos: list[tuple[str, bytes]] = []
    for q_key, label in view_keys:
        fname = q_map.get(q_key, "").strip()
        if not fname:
            continue
        if fname in attachments:
            labeled_photos.append((label, attachments[fname]))
        elif fname.lower() in lower_attachments:
            _, img_bytes = lower_attachments[fname.lower()]
            labeled_photos.append((label, img_bytes))

    return {
        "submission_id": submission_id,
        "hydrant_number": hydrant_number,
        "timestamp": timestamp,
        "labeled_photos": labeled_photos,
    }


# ─────────────────────────────────────────────────────────────────────────────
# XLSX PARSING
# ─────────────────────────────────────────────────────────────────────────────

def parse_xlsx(xlsx_bytes: bytes) -> dict[str, dict]:
    """
    Parse the submission log Excel file.
    Returns {submission_id: row_dict}.
    Raises ValueError if Submission ID column not found.
    """
    df = pd.read_excel(BytesIO(xlsx_bytes))
    df.columns = [str(c).strip() for c in df.columns]
    norm = {c.lower(): c for c in df.columns}

    id_col = next(
        (norm[k] for k in norm if "submission id" in k or "submission_id" in k),
        None,
    )
    if id_col is None:
        raise ValueError(
            "Could not find a 'Submission ID' column in the Excel file. "
            f"Found columns: {list(df.columns)}"
        )

    result: dict[str, dict] = {}
    for _, row in df.iterrows():
        sid = str(row[id_col]).strip()
        if sid and sid.lower() != "nan":
            result[sid] = row.to_dict()
    return result


def parse_contractor_xlsx(xlsx_bytes: bytes) -> dict[tuple, dict]:
    """
    Parse the contractor submission log Excel file.
    Returns {(submission_id, hydrant_number): row_dict} for dual-key matching.
    Raises ValueError if Submission ID column not found.
    """
    df = pd.read_excel(BytesIO(xlsx_bytes))
    df.columns = [str(c).strip() for c in df.columns]
    norm = {c.lower(): c for c in df.columns}

    id_col = next(
        (norm[k] for k in norm if "submission id" in k or "submission_id" in k),
        None,
    )
    if id_col is None:
        raise ValueError(
            "Could not find a 'Submission ID' column in the contractor Excel file. "
            f"Found columns: {list(df.columns)}"
        )

    hydrant_num_col  = next((norm[k] for k in norm if "hydrant number" in k), None)
    fire_station_col = next((norm[k] for k in norm if "fire station" in k), None)
    timestamp_col    = next((norm[k] for k in norm if "submission time" in k or k == "timestamp"), None)
    street_col       = next((norm[k] for k in norm if "street name" in k), None)
    hydrant_type_col = next((norm[k] for k in norm if "hydrant type" in k), None)
    defects_col      = next((norm[k] for k in norm if "list of defects" in k or k == "defects"), None)

    def _get(row, col):
        if col is None:
            return ""
        val = str(row.get(col, "")).strip()
        return "" if val.lower() == "nan" else val

    result: dict[tuple, dict] = {}
    for _, row in df.iterrows():
        sid  = str(row[id_col]).strip()
        if not sid or sid.lower() == "nan":
            continue
        hnum = _get(row, hydrant_num_col)
        key  = (sid, hnum)
        result[key] = {
            "submission_id":   sid,
            "hydrant_number":  hnum,
            "fire_station":    _get(row, fire_station_col),
            "submission_time": _get(row, timestamp_col),
            "street_name":     _get(row, street_col),
            "hydrant_type":    _get(row, hydrant_type_col),
            "defects":         _get(row, defects_col),
        }
    return result


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL REPORT GENERATION
# ─────────────────────────────────────────────────────────────────────────────

def generate_excel(entries: list[dict]) -> bytes:
    """Generate a rectification report Excel workbook and return as bytes.

    Layout per entry:
      Row N   — text data (Submission ID, Fire Station, Street Name, Hydrant, Defects)
      Row N+1 — embedded defect photos (one per column, up to 4), if any exist
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Rectification Required"

    # ── Header row ────────────────────────────────────────────────────────────
    headers = ["Submission ID", "Fire Station", "Street Name", "Hydrant", "List of Defects"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # ── Data rows ─────────────────────────────────────────────────────────────
    IMG_SIZE = 200       # max px per side for thumbnail
    IMG_PT   = 155       # row height in points (~200 px)
    IMG_COL_W = 28       # column width in Excel units (~200 px)

    current_row = 2
    for entry in entries:
        hydrant = f"{entry.get('hydrant_type', '')} {entry.get('hydrant_number', '')}".strip()
        ws.cell(row=current_row, column=1, value=entry.get("submission_id", ""))
        ws.cell(row=current_row, column=2, value=entry.get("fire_station", ""))
        ws.cell(row=current_row, column=3, value=entry.get("street_name", ""))
        ws.cell(row=current_row, column=4, value=hydrant)
        ws.cell(row=current_row, column=5, value=entry.get("defects", ""))

        photos: list[tuple[str, bytes]] = entry.get("photos", [])
        if photos:
            photo_row = current_row + 1
            ws.row_dimensions[photo_row].height = IMG_PT

            for col_idx, (fname, img_bytes) in enumerate(photos[:4], 1):
                try:
                    img = Image.open(BytesIO(img_bytes))
                    img.thumbnail((IMG_SIZE, IMG_SIZE))
                    # Ensure RGB so PNG save always works
                    if img.mode not in ("RGB", "RGBA"):
                        img = img.convert("RGB")
                    out = BytesIO()
                    img.save(out, format="PNG")
                    out.seek(0)

                    xl_img = XLImage(out)
                    xl_img.width = IMG_SIZE
                    xl_img.height = IMG_SIZE
                    ws.add_image(xl_img, f"{get_column_letter(col_idx)}{photo_row}")

                    # Widen column to fit image (only expand, never shrink)
                    col_letter = get_column_letter(col_idx)
                    if ws.column_dimensions[col_letter].width < IMG_COL_W:
                        ws.column_dimensions[col_letter].width = IMG_COL_W
                except Exception:
                    pass  # skip unreadable images silently

            current_row += 2  # data row + photo row
        else:
            current_row += 1

    # ── Auto-width for text columns (A–E), skip if already widened for photos ─
    for col_idx in range(1, 6):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=0,
        )
        desired = min(max_len + 4, 50)
        if ws.column_dimensions[col_letter].width < desired:
            ws.column_dimensions[col_letter].width = desired

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_contractor_excel(entries: list[dict], sheet_title: str) -> bytes:
    """Generate a contractor rectification report Excel workbook.

    Layout per entry:
      Row N   — text data (Fire Station, Submission Time, Submission ID, Street Name, Hydrant, Defects[, Reason])
      Row N+1 — photo labels (Front View, Left Side, Right Side, Back View)
      Row N+2 — embedded photos (one per column, up to 4)
    If no photos, entry occupies 1 row only.
    Includes a 'Reason for Not Satisfactory' column when any entry has a reason.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel sheet name limit

    has_reason = any(e.get("reason", "").strip() for e in entries)
    headers = ["Fire Station", "Submission Time", "Submission ID", "Street Name", "Hydrant", "List of Defects"]
    if has_reason:
        headers.append("Reason for Not Satisfactory")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    IMG_SIZE  = 200
    IMG_PT    = 155
    IMG_COL_W = 28

    current_row = 2
    for entry in entries:
        hydrant = f"{entry.get('hydrant_type', '')} {entry.get('hydrant_number', '')}".strip()
        ws.cell(row=current_row, column=1, value=entry.get("fire_station", ""))
        ws.cell(row=current_row, column=2, value=entry.get("submission_time", ""))
        ws.cell(row=current_row, column=3, value=entry.get("submission_id", ""))
        ws.cell(row=current_row, column=4, value=entry.get("street_name", ""))
        ws.cell(row=current_row, column=5, value=hydrant)
        ws.cell(row=current_row, column=6, value=entry.get("defects", ""))
        if has_reason:
            ws.cell(row=current_row, column=7, value=entry.get("reason", ""))

        labeled_photos: list[tuple[str, bytes]] = entry.get("labeled_photos", [])
        if labeled_photos:
            label_row = current_row + 1
            photo_row = current_row + 2
            ws.row_dimensions[photo_row].height = IMG_PT

            for col_idx, (label, img_bytes) in enumerate(labeled_photos[:4], 1):
                # Write label text
                ws.cell(row=label_row, column=col_idx, value=label)

                # Embed image
                try:
                    img = Image.open(BytesIO(img_bytes))
                    img.thumbnail((IMG_SIZE, IMG_SIZE))
                    if img.mode not in ("RGB", "RGBA"):
                        img = img.convert("RGB")
                    out = BytesIO()
                    img.save(out, format="PNG")
                    out.seek(0)

                    xl_img = XLImage(out)
                    xl_img.width = IMG_SIZE
                    xl_img.height = IMG_SIZE
                    ws.add_image(xl_img, f"{get_column_letter(col_idx)}{photo_row}")

                    col_letter = get_column_letter(col_idx)
                    if ws.column_dimensions[col_letter].width < IMG_COL_W:
                        ws.column_dimensions[col_letter].width = IMG_COL_W
                except Exception:
                    pass

            current_row += 3  # data + label + photo rows
        else:
            current_row += 1

    # Auto-width for text columns
    num_text_cols = 7 if has_reason else 6
    for col_idx in range(1, num_text_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=0,
        )
        desired = min(max_len + 4, 50)
        if ws.column_dimensions[col_letter].width < desired:
            ws.column_dimensions[col_letter].width = desired

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# LOGIN
# ─────────────────────────────────────────────────────────────────────────────

def show_login():
    col_l, col_c, col_r = st.columns([1, 2, 1])
    with col_c:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown("## 🔒 SCDF Hydrant Form Checking")
        st.markdown("Please enter the password to continue.")

        with st.form("login_form"):
            password = st.text_input("Password", type="password", placeholder="Enter password")
            submitted = st.form_submit_button("Login", use_container_width=True, type="primary")
        if submitted:
            if password == APP_PASSWORD:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Incorrect password. Please try again.")

    st.stop()


# ─────────────────────────────────────────────────────────────────────────────
# WSN HYDRANT CHECKS SECTION
# ─────────────────────────────────────────────────────────────────────────────

def show_wsn_section():
    # ── Upload zone ───────────────────────────────────────────────────────────
    col1, col2 = st.columns(2)
    with col1:
        xlsx_file = st.file_uploader(
            "Submission Log (.xlsx)",
            type=["xlsx"],
            help="The Excel file exported from the hydrant inspection system.",
            key="wsn_xlsx_upload",
        )
    with col2:
        eml_files = st.file_uploader(
            f"FormSG Emails (.eml) — max {MAX_UPLOAD_MB} MB total",
            type=["eml"],
            accept_multiple_files=True,
            help="One or more .eml files exported from Outlook.",
            key="wsn_eml_upload",
        )

    # ── Process ───────────────────────────────────────────────────────────────
    if xlsx_file and eml_files:
        # Check total EML size
        total_mb = sum(f.size for f in eml_files) / (1024 * 1024)
        if total_mb > MAX_UPLOAD_MB:
            st.error(f"Total EML size ({total_mb:.1f} MB) exceeds the {MAX_UPLOAD_MB} MB limit.")
            return

        if st.button("Process Files", type="primary", use_container_width=False, key="wsn_process_btn"):
            with st.spinner("Parsing files and matching entries…"):
                # Parse XLSX
                try:
                    xlsx_data = parse_xlsx(xlsx_file.read())
                except ValueError as e:
                    st.error(str(e))
                    return

                # Parse each EML
                eml_data: dict[str, dict] = {}
                parse_errors: list[str] = []
                for eml_f in eml_files:
                    entry = parse_eml(eml_f.read())
                    if entry.get("submission_id"):
                        eml_data[entry["submission_id"]] = entry
                    else:
                        parse_errors.append(eml_f.name)

                if parse_errors:
                    st.warning(
                        f"Could not extract form data from: {', '.join(parse_errors)}. "
                        "These files were skipped."
                    )

                # Match
                matched_ids = set(xlsx_data.keys()) & set(eml_data.keys())
                matched_entries = [eml_data[sid] for sid in sorted(matched_ids)]

                if not matched_entries:
                    st.warning(
                        "No matching Submission IDs found between the Excel file and the EML files. "
                        "Please verify the files are from the same submission batch."
                    )
                    return

                st.session_state.matched_entries = matched_entries
                st.session_state.indications = {e["submission_id"]: None for e in matched_entries}
                st.rerun()

    # ── Results ───────────────────────────────────────────────────────────────
    if "matched_entries" not in st.session_state:
        return

    entries: list[dict] = st.session_state.matched_entries
    indications: dict = st.session_state.indications

    st.markdown(f"### {len(entries)} matched {'entry' if len(entries) == 1 else 'entries'}")
    st.markdown("Review each entry and select an indication before generating the report.")
    st.divider()

    for entry in entries:
        sid = entry["submission_id"]
        hydrant = f"{entry.get('hydrant_type', '')} {entry.get('hydrant_number', '')}".strip()

        with st.container(border=True):
            # ── Header row ────────────────────────────────────────────────────
            h1, h2, h3 = st.columns([2, 2, 2])
            h1.markdown(f"**Submission ID**  \n{sid}")
            h2.markdown(f"**Street Name**  \n{entry.get('street_name', '—')}")
            h3.markdown(f"**Hydrant**  \n{hydrant or '—'}")

            # ── Defects ───────────────────────────────────────────────────────
            st.markdown(f"**List of Defects:** {entry.get('defects', '—')}")

            # ── Photos ────────────────────────────────────────────────────────
            photos: list[tuple[str, bytes]] = entry.get("photos", [])
            if photos:
                num_cols = min(len(photos), 4)
                photo_cols = st.columns(num_cols)
                for i, (fname, img_bytes) in enumerate(photos):
                    with photo_cols[i % num_cols]:
                        try:
                            img = Image.open(BytesIO(img_bytes))
                            st.image(img, caption=fname, use_container_width=True)
                        except Exception:
                            st.warning(f"Could not display image: {fname}")
            else:
                st.caption("No defect photos attached.")

            # ── Indication ────────────────────────────────────────────────────
            choice = st.radio(
                "Indication",
                ["Requires Rectification", "Does Not Require Rectification"],
                key=f"ind_{sid}",
                index=None,
                horizontal=True,
            )
            # Mirror choice into indications dict (radio key keeps it in session state
            # but we also need it accessible for the generate step below)
            indications[sid] = st.session_state.get(f"ind_{sid}")

    # ── Generate Excel ────────────────────────────────────────────────────────
    st.divider()
    all_answered = all(
        st.session_state.get(f"ind_{e['submission_id']}") is not None
        for e in entries
    )

    if not all_answered:
        unanswered = sum(
            1 for e in entries
            if st.session_state.get(f"ind_{e['submission_id']}") is None
        )
        st.info(f"{unanswered} {'entry has' if unanswered == 1 else 'entries have'} not been vetted yet.")
    else:
        rectification_entries = [
            e for e in entries
            if st.session_state.get(f"ind_{e['submission_id']}") == "Requires Rectification"
        ]

        if not rectification_entries:
            st.success("All entries vetted — no rectification required.")
        else:
            st.markdown(
                f"**{len(rectification_entries)} of {len(entries)} entries** require rectification."
            )
            if st.button("Generate Excel Report", type="primary", key="wsn_generate_btn"):
                excel_bytes = generate_excel(rectification_entries)
                st.download_button(
                    label="⬇️ Download Rectification Report",
                    data=excel_bytes,
                    file_name="rectification_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=False,
                )


# ─────────────────────────────────────────────────────────────────────────────
# CONTRACTOR RECTIFICATION CHECKS SECTION
# ─────────────────────────────────────────────────────────────────────────────

def show_contractor_section():
    # ── Upload zone ───────────────────────────────────────────────────────────
    col1, col2 = st.columns(2)
    with col1:
        cr_xlsx = st.file_uploader(
            "Contractor Submission Log (.xlsx)",
            type=["xlsx"],
            help="The Excel file with contractor submissions.",
            key="cr_xlsx_upload",
        )
    with col2:
        cr_emls = st.file_uploader(
            f"Contractor FormSG Emails (.eml) — max {MAX_UPLOAD_MB} MB total",
            type=["eml"],
            accept_multiple_files=True,
            help="One or more contractor .eml files exported from Outlook.",
            key="cr_eml_upload",
        )

    # ── Process ───────────────────────────────────────────────────────────────
    if cr_xlsx and cr_emls:
        total_mb = sum(f.size for f in cr_emls) / (1024 * 1024)
        if total_mb > MAX_UPLOAD_MB:
            st.error(f"Total EML size ({total_mb:.1f} MB) exceeds the {MAX_UPLOAD_MB} MB limit.")
            return

        if st.button("Process Contractor Files", type="primary", key="cr_process_btn"):
            with st.spinner("Parsing files and matching entries…"):
                # Parse XLSX
                try:
                    cr_xlsx_data = parse_contractor_xlsx(cr_xlsx.read())
                except ValueError as e:
                    st.error(str(e))
                    return

                # Parse each EML
                cr_eml_data: dict[tuple, dict] = {}
                parse_errors: list[str] = []
                for eml_f in cr_emls:
                    entry = parse_contractor_eml(eml_f.read())
                    if entry.get("submission_id"):
                        key = (entry["submission_id"], entry["hydrant_number"])
                        cr_eml_data[key] = entry
                    else:
                        parse_errors.append(eml_f.name)

                if parse_errors:
                    st.warning(
                        f"Could not extract form data from: {', '.join(parse_errors)}. "
                        "These files were skipped."
                    )

                # Dual-key match: (submission_id, hydrant_number) must match in both
                matched_keys = set(cr_xlsx_data.keys()) & set(cr_eml_data.keys())
                matched_entries = []
                for key in sorted(matched_keys):
                    xlsx_row = cr_xlsx_data[key]
                    eml_entry = cr_eml_data[key]
                    merged = {**xlsx_row, "labeled_photos": eml_entry.get("labeled_photos", [])}
                    matched_entries.append(merged)

                if not matched_entries:
                    st.warning(
                        "No matching entries found. Both Submission ID and Hydrant Number "
                        "must match between the Excel file and the EML files."
                    )
                    return

                st.session_state.cr_matched_entries = matched_entries
                st.rerun()

    # ── Results ───────────────────────────────────────────────────────────────
    if "cr_matched_entries" not in st.session_state:
        return

    entries: list[dict] = st.session_state.cr_matched_entries

    st.markdown(f"### {len(entries)} matched {'entry' if len(entries) == 1 else 'entries'}")
    st.markdown("Review each entry and indicate whether the contractor's work is satisfactory.")
    st.divider()

    for entry in entries:
        sid    = entry["submission_id"]
        hnum   = entry.get("hydrant_number", "")
        hydrant = f"{entry.get('hydrant_type', '')} {hnum}".strip()

        with st.container(border=True):
            h1, h2, h3 = st.columns([2, 2, 2])
            h1.markdown(f"**Submission ID**  \n{sid}")
            h2.markdown(f"**Street Name**  \n{entry.get('street_name', '—')}")
            h3.markdown(f"**Hydrant**  \n{hydrant or hnum or '—'}")

            st.markdown(f"**List of Defects:** {entry.get('defects', '—')}")

            labeled_photos: list[tuple[str, bytes]] = entry.get("labeled_photos", [])
            if labeled_photos:
                photo_cols = st.columns(4)
                for i, (label, img_bytes) in enumerate(labeled_photos):
                    with photo_cols[i]:
                        try:
                            img = Image.open(BytesIO(img_bytes))
                            st.image(img, caption=label, use_container_width=True)
                        except Exception:
                            st.warning(f"Could not display: {label}")
            else:
                st.caption("No photos attached.")

            st.radio(
                "Indication",
                ["Satisfactory", "Not Satisfactory"],
                key=f"cr_ind_{sid}",
                index=None,
                horizontal=True,
            )

            # Mandatory reason field for not satisfactory entries
            if st.session_state.get(f"cr_ind_{sid}") == "Not Satisfactory":
                st.text_area(
                    "Reason for Not Satisfactory (mandatory)",
                    key=f"cr_reason_{sid}",
                    placeholder="Describe why the work is not satisfactory…",
                    height=80,
                )

    # ── Generate Reports ──────────────────────────────────────────────────────
    st.divider()
    all_answered = all(
        st.session_state.get(f"cr_ind_{e['submission_id']}") is not None
        for e in entries
    )
    all_reasons_filled = all(
        st.session_state.get(f"cr_reason_{e['submission_id']}", "").strip() != ""
        for e in entries
        if st.session_state.get(f"cr_ind_{e['submission_id']}") == "Not Satisfactory"
    )

    if not all_answered:
        unanswered = sum(
            1 for e in entries
            if st.session_state.get(f"cr_ind_{e['submission_id']}") is None
        )
        st.info(f"{unanswered} {'entry has' if unanswered == 1 else 'entries have'} not been indicated yet.")
    elif not all_reasons_filled:
        missing = sum(
            1 for e in entries
            if st.session_state.get(f"cr_ind_{e['submission_id']}") == "Not Satisfactory"
            and not st.session_state.get(f"cr_reason_{e['submission_id']}", "").strip()
        )
        st.warning(
            f"{missing} not satisfactory {'entry is' if missing == 1 else 'entries are'} "
            "missing a reason. Please fill in all reason fields before generating reports."
        )
    else:
        satisfactory_entries = [
            e for e in entries
            if st.session_state.get(f"cr_ind_{e['submission_id']}") == "Satisfactory"
        ]
        not_satisfactory_entries = [
            e for e in entries
            if st.session_state.get(f"cr_ind_{e['submission_id']}") == "Not Satisfactory"
        ]

        XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        # ── Satisfactory report ───────────────────────────────────────────────
        st.markdown("#### Satisfactory Report")
        if not satisfactory_entries:
            st.info("No satisfactory entries.")
        else:
            st.markdown(
                f"**{len(satisfactory_entries)}** satisfactory "
                f"{'entry' if len(satisfactory_entries) == 1 else 'entries'}."
            )
            sat_format = st.radio(
                "Report format",
                ["Compiled Report", "Separated by Fire Station"],
                key="cr_sat_format",
                horizontal=True,
            )

            if sat_format == "Compiled Report":
                if st.button("Generate Satisfactory Report", type="primary", key="cr_gen_sat_btn"):
                    st.session_state.cr_sat_compiled = generate_contractor_excel(
                        satisfactory_entries, "Satisfactory"
                    )
                    st.session_state.cr_sat_sep_reports = {}
                if st.session_state.get("cr_sat_compiled"):
                    st.download_button(
                        label="⬇️ Download Satisfactory Report",
                        data=st.session_state.cr_sat_compiled,
                        file_name="satisfactory_report.xlsx",
                        mime=XLSX_MIME,
                        key="cr_dl_sat_compiled",
                    )

            else:  # Separated by Fire Station
                from collections import defaultdict
                grouped: dict[str, list] = defaultdict(list)
                for e in satisfactory_entries:
                    grouped[e.get("fire_station", "Unknown")].append(e)

                if st.button("Generate Reports by Fire Station", type="primary", key="cr_gen_sat_sep_btn"):
                    st.session_state.cr_sat_sep_reports = {
                        fs: generate_contractor_excel(group, fs)
                        for fs, group in sorted(grouped.items())
                    }
                    st.session_state.cr_sat_compiled = None

                for fs, report_bytes in st.session_state.get("cr_sat_sep_reports", {}).items():
                    safe_name = fs.replace("/", "-").replace("\\", "-")
                    st.download_button(
                        label=f"⬇️ Download — {fs}",
                        data=report_bytes,
                        file_name=f"satisfactory_{safe_name}.xlsx",
                        mime=XLSX_MIME,
                        key=f"cr_dl_sat_{fs}",
                    )

        st.markdown("---")

        # ── Not Satisfactory report ───────────────────────────────────────────
        st.markdown("#### Not Satisfactory Report")
        if not not_satisfactory_entries:
            st.info("No unsatisfactory entries.")
        else:
            st.markdown(
                f"**{len(not_satisfactory_entries)}** not satisfactory "
                f"{'entry' if len(not_satisfactory_entries) == 1 else 'entries'}."
            )
            if st.button("Generate Not Satisfactory Report", type="primary", key="cr_gen_notsat_btn"):
                not_sat_with_reasons = []
                for e in not_satisfactory_entries:
                    entry_copy = dict(e)
                    entry_copy["reason"] = st.session_state.get(
                        f"cr_reason_{e['submission_id']}", ""
                    )
                    not_sat_with_reasons.append(entry_copy)
                st.session_state.cr_notsat_report = generate_contractor_excel(
                    not_sat_with_reasons, "Not Satisfactory"
                )
            if st.session_state.get("cr_notsat_report"):
                st.download_button(
                    label="⬇️ Download Not Satisfactory Report",
                    data=st.session_state.cr_notsat_report,
                    file_name="not_satisfactory_report.xlsx",
                    mime=XLSX_MIME,
                    key="cr_dl_notsat_btn",
                )


# ─────────────────────────────────────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────────────────────────────────────

def show_app():
    st.title("🚒 SCDF Hydrant Form Checking")
    st.caption("Upload files and vet submissions.")

    tab1, tab2 = st.tabs(["SCDF Form Submission Checks", "Contractor Rectification Checks"])
    with tab1:
        show_wsn_section()
    with tab2:
        show_contractor_section()


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if not st.session_state.get("authenticated"):
    show_login()
else:
    show_app()
